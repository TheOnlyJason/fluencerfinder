"""Import influencers from the multi-sheet "Influencer stuff.xlsx" workbook.

Only rows whose name/handle cell contains a hyperlink to a supported platform
(Instagram, TikTok, X, YouTube, Twitch) are imported — the link is the only
reliable source of platform + exact handle. Rows without a link, on unsupported
platforms (Facebook), or marked "blacklist" are skipped.

Tier band -> follower_count uses the band midpoint. GENRE -> niche.

Usage:
    python scripts/import_xlsx.py "Influencer stuff.xlsx"           # apply
    python scripts/import_xlsx.py "Influencer stuff.xlsx" --dry-run # report only
"""
import argparse
import re
import sys
import warnings
from datetime import datetime
from typing import Optional

import openpyxl
from sqlmodel import Session, select

sys.path.insert(0, ".")

from app.db.session import engine  # noqa: E402
from app.models.account import Account  # noqa: E402
from app.models.enums import Platform  # noqa: E402
from app.utils.handles import build_profile_url, normalize_handle, parse_profile_url  # noqa: E402
from app.utils.tiers import tier_midpoint  # noqa: E402

warnings.simplefilter("ignore")

INFLUENCER_SHEETS = ["GMK", "RUSA", "REU", "RAS", "RCN", "LFD", "RS"]
NAME_COL_CANDIDATES = ("NAME", "HANDLE", "upl")
SKIP_STATUSES = {"blacklist"}

_TIER_RE = re.compile(r"tier\s*([1-5])", re.I)


def parse_tier(value) -> Optional[int]:
    if value is None:
        return None
    m = _TIER_RE.search(str(value))
    return int(m.group(1)) if m else None


def clean_text(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).replace("\n", " ").strip()
    # Drop a trailing "handle / Display Name" or "handle; Display Name" combo's
    # display portion is handled by the caller; here just normalize whitespace.
    return text or None


def header_map(ws):
    out = {}
    for cell in ws[1]:
        if cell.value is not None:
            out[str(cell.value).strip().lower()] = cell.column
    return out


def find_col(headers, candidates):
    for cand in candidates:
        if cand.lower() in headers:
            return headers[cand.lower()]
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("path")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.path)

    # Collect best record per (platform, handle) across all sheets.
    records: dict[tuple[str, str], dict] = {}
    stats = {"rows": 0, "linked": 0, "unsupported": 0, "blacklist": 0, "no_handle": 0}
    by_platform: dict[str, int] = {}

    for sheet in INFLUENCER_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        headers = header_map(ws)
        name_col = find_col(headers, NAME_COL_CANDIDATES)
        if not name_col:
            continue
        genre_col = find_col(headers, ("GENRE", "genre"))
        tier_col = find_col(headers, ("TIER", "tier"))
        status_col = find_col(headers, ("STATUS", "status"))
        loc_col = find_col(headers, ("Location", "location"))

        for row in ws.iter_rows(min_row=2):
            cell = row[name_col - 1]
            if cell.value is None and not cell.hyperlink:
                continue
            stats["rows"] += 1

            status = clean_text(row[status_col - 1].value) if status_col else None
            if status and status.lower() in SKIP_STATUSES:
                stats["blacklist"] += 1
                continue

            url = cell.hyperlink.target if cell.hyperlink else None
            if not url:
                continue
            platform, handle = parse_profile_url(url)
            if not platform:
                stats["unsupported"] += 1
                continue
            handle = normalize_handle(handle or "")
            if not handle:
                stats["no_handle"] += 1
                continue
            stats["linked"] += 1

            display_text = clean_text(cell.value)
            # Keep display name only when it adds info beyond the handle.
            display_name = None
            if display_text:
                # "handle / Name" or "handle; Name" -> take the descriptive part
                parts = re.split(r"\s*[/;]\s*", display_text)
                cand = parts[-1].strip() if len(parts) > 1 else display_text
                if normalize_handle(cand) != handle:
                    display_name = cand

            tier = parse_tier(row[tier_col - 1].value) if tier_col else None
            follower_count = tier_midpoint(tier) if tier else None
            niche = clean_text(row[genre_col - 1].value) if genre_col else None
            if niche:
                niche = niche.lower()
            location = clean_text(row[loc_col - 1].value) if loc_col else None

            key = (platform.value, handle)
            existing = records.get(key)
            rec = {
                "platform": platform,
                "handle": handle,
                "display_name": display_name,
                "profile_url": url.strip(),
                "follower_count": follower_count,
                "niche": niche,
                "channel_type": niche,
                "location_text": location,
            }
            if existing is None:
                records[key] = rec
                by_platform[platform.value] = by_platform.get(platform.value, 0) + 1
            else:
                # Merge: fill gaps, keep the higher follower estimate.
                for f in ("display_name", "niche", "channel_type", "location_text"):
                    if not existing.get(f) and rec.get(f):
                        existing[f] = rec[f]
                if rec["follower_count"] and (
                    not existing["follower_count"]
                    or rec["follower_count"] > existing["follower_count"]
                ):
                    existing["follower_count"] = rec["follower_count"]

    print("Scan stats:", stats)
    print("Unique by platform:", by_platform)
    print("Total unique importable:", len(records))

    if args.dry_run:
        print("\n[dry-run] no DB writes. Sample:")
        for rec in list(records.values())[:10]:
            print(" ", rec["platform"].value, rec["handle"], "| f=", rec["follower_count"],
                  "| niche=", rec["niche"], "| name=", rec["display_name"])
        return

    inserted = updated = 0
    with Session(engine) as session:
        for rec in records.values():
            existing = session.exec(
                select(Account).where(
                    Account.platform == rec["platform"],
                    Account.handle == rec["handle"],
                )
            ).first()
            now = datetime.utcnow()
            if existing:
                if not existing.follower_count and rec["follower_count"]:
                    existing.follower_count = rec["follower_count"]
                if not existing.niche and rec["niche"]:
                    existing.niche = rec["niche"]
                    existing.channel_type = existing.channel_type or rec["channel_type"]
                if not existing.display_name and rec["display_name"]:
                    existing.display_name = rec["display_name"]
                if not existing.location_text and rec["location_text"]:
                    existing.location_text = rec["location_text"]
                existing.updated_at = now
                existing.is_active = True
                session.add(existing)
                updated += 1
            else:
                session.add(Account(
                    platform=rec["platform"],
                    handle=rec["handle"],
                    display_name=rec["display_name"],
                    profile_url=rec["profile_url"] or build_profile_url(rec["platform"], rec["handle"]),
                    niche=rec["niche"],
                    channel_type=rec["channel_type"],
                    location_text=rec["location_text"],
                    follower_count=rec["follower_count"],
                    is_active=True,
                ))
                inserted += 1
            if (inserted + updated) % 500 == 0:
                session.commit()
        session.commit()

    print(f"\nDone. Inserted={inserted} Updated={updated}")


if __name__ == "__main__":
    main()
